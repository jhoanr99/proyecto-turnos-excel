# pip install flask openpyxl google-cloud-firestore firebase-admin
import io
from datetime import datetime, timedelta
# !! IMPORTANTE !! Añadir render_template
from flask import Flask, request, send_file, jsonify, render_template
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl.utils import range_boundaries

import firebase_admin
from firebase_admin import credentials, firestore
from google.cloud.firestore_v1 import FieldFilter

# ------------ Config ------------
JSON_PATH = 'turnosd1app-firebase-adminsdk-fbsvc-beb0f93c56.json'
TEMPLATE  = 'programacion-y-control-de-horas-actualizado.xlsx'

# Inicializar Firebase
if not firebase_admin._apps:
    cred = credentials.Certificate(JSON_PATH)
    firebase_admin.initialize_app(cred)
db = firestore.client()

# ... (El resto de tus funciones auxiliares y generar_excel se mantienen igual) ...
def top_left_of_merge(ws, row, col):
    for mr in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(mr))
        if min_r <= row <= max_r and min_c <= col <= max_c:
            return (min_r, min_c)
    return (row, col)

def set_cell_value_safe(ws, row, col, value):
    r0, c0 = top_left_of_merge(ws, row, col)
    ws.cell(row=r0, column=c0).value = value

def clear_cells_but_skip_merged(ws, start_row=11, start_col=3):
    merged_cells_coords = set()
    for mr in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(mr))
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                merged_cells_coords.add((r, c))

    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, ws.max_column + 1):
            if (r, c) not in merged_cells_coords:
                ws.cell(row=r, column=c).value = None

def to_dt(v):
    if isinstance(v, datetime):
        return v
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(str(v), fmt)
        except (ValueError, TypeError):
            continue
    return None

def to_bool(v):
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ('true', '1', 'si', 'sí', 'y', 'yes', 't', 'on')

def pick(turno, *keys):
    for k in keys:
        if k in turno and turno.get(k) is not None and str(turno.get(k, '')).strip() != '':
            return turno.get(k)
    return ''

def generar_excel(fecha_ini: datetime, fecha_fin: datetime):
    # 1. Obtener datos de Firestore
    fechas_str = [(fecha_ini + timedelta(days=d)).strftime('%d/%m/%Y')
                  for d in range((fecha_fin - fecha_ini).days + 1)]
    docs = db.collection('turnos').where(filter=FieldFilter('fecha', 'in', fechas_str)).stream()

    turnos = []
    for doc in docs:
        d = doc.to_dict()
        d['id'] = doc.id
        dt = to_dt(d.get('fecha'))
        if not dt:
            continue
        d['FECHA'] = dt
        d['nombre'] = str(d.get('nombre', '')).strip()
        d['cargo']  = str(d.get('cargo', '')).strip()
        turnos.append(d)

    # 2. Cargar la plantilla de Excel
    wb = load_workbook(TEMPLATE)
    ws = wb['Hoja1']

    # 3. Limpiar el área de datos existente
    clear_cells_but_skip_merged(ws, start_row=11, start_col=3)

    # 4. Escribir encabezados
    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    ws.cell(row=5, column=16).value = meses[fecha_ini.month - 1]
    texto_semana = f"del {fecha_ini.day} al {fecha_fin.day}"
    ws.cell(row=5, column=24).value = texto_semana
    HEADER_ROW = 6
    dia_a_columna = {0: 3, 1: 7, 2: 11, 3: 15, 4: 19, 5: 23, 6: 27}

    fechas_semana = sorted({t['FECHA'].date() for t in turnos})
    for f in fechas_semana:
        col = dia_a_columna.get(f.weekday())
        if col:
            set_cell_value_safe(ws, HEADER_ROW, col, f.day)

    # 5. Definir el orden de los usuarios
    usuarios = defaultdict(lambda: {'CARGO': ''})
    for t in turnos:
        nom = t['nombre']
        car = t['cargo']
        if nom:
            usuarios[nom]['CARGO'] = car or usuarios[nom]['CARGO']

    cargo_orden = {'supervisor': 0, 'full-time': 1, 'part-time': 2}
    orden = sorted(
        usuarios.keys(),
        key=lambda n: (cargo_orden.get(usuarios[n]['CARGO'].lower(), 3), n.lower())
    )
    
    start_row = 11
    row_step  = 5
    user_rows = {n.lower(): start_row + i * row_step for i, n in enumerate(orden)}

    # 6. Agrupar los turnos por empleado
    por_emp = defaultdict(list)
    for t in turnos:
        por_emp[t['nombre'].lower()].append(t)
    for k in por_emp:
        por_emp[k].sort(key=lambda x: x['FECHA'])

    # 7. Función para escribir un turno
    def escribir_turno_exact(fila_base, turno, col_entrada1):
        usa_partido = to_bool(pick(turno, 'usaTurnoPartido', 'USATURNOPARTIDO'))
        
        receso_val = str(pick(turno, 'receso', 'HORA RECESO', 'HORA_RECESO', 'horaReceso')).strip()
        if receso_val.lower() == '30 minutos':
            receso = '30'
        elif receso_val.lower() == 'no aplica':
            receso = ''
        else:
            receso = receso_val

        if usa_partido:
            entrada1 = str(pick(turno, 'entrada1', 'HORA ENTRADA 1', 'HORA_ENTRADA_1', 'horaEntrada1', 'ENTRADA1') or '').strip()
            salida1  = str(pick(turno, 'salida1', 'HORA SALIDA 1', 'HORA_SALIDA_1', 'horaSalida1', 'SALIDA1') or '').strip()
            entrada2 = str(pick(turno, 'entrada2', 'HORA ENTRADA 2', 'HORA_ENTRADA_2', 'horaEntrada2', 'ENTRADA2') or '').strip()
            salida2  = str(pick(turno, 'salida2', 'HORA SALIDA 2', 'HORA_SALIDA_2', 'horaSalida2', 'SALIDA2') or '').strip()

            set_cell_value_safe(ws, fila_base, col_entrada1, entrada1)
            set_cell_value_safe(ws, fila_base, col_entrada1 + 1, receso)
            set_cell_value_safe(ws, fila_base, col_entrada1 + 2, salida1)
            
            if entrada2 or salida2:
                set_cell_value_safe(ws, fila_base + 1, col_entrada1, entrada2)
                set_cell_value_safe(ws, fila_base + 1, col_entrada1 + 2, salida2)
        else:
            entrada1 = str(pick(turno, 'entrada', 'ENTRADA', 'HORA_ENTRADA', 'HORA ENTRADA', 'horaEntrada') or '').strip()
            salida1  = str(pick(turno, 'salida', 'SALIDA', 'HORA_SALIDA', 'HORA SALIDA', 'horaSalida') or '').strip()
            
            set_cell_value_safe(ws, fila_base, col_entrada1, entrada1)
            set_cell_value_safe(ws, fila_base, col_entrada1 + 1, receso)
            set_cell_value_safe(ws, fila_base, col_entrada1 + 2, salida1)
            
        estado = str(pick(turno, 'estado', 'ESTADO')).strip()
        if estado:
            set_cell_value_safe(ws, fila_base + 2, col_entrada1, estado)

    # 8. Bucle principal para escribir nombres y turnos
    for nombre_key, fila_base in user_rows.items():
        candidates = [n for n in orden if n.lower() == nombre_key]
        original_name = candidates[0] if candidates else nombre_key
        set_cell_value_safe(ws, fila_base, 2, original_name)

        cargo = usuarios.get(original_name, {}).get('CARGO', '')
        if cargo:
            set_cell_value_safe(ws, fila_base + 1, 2, cargo)

        for t in por_emp.get(nombre_key, []):
            dia = t['FECHA'].weekday()
            col = dia_a_columna.get(dia)
            if not col:
                continue
            escribir_turno_exact(fila_base, t, col)

    # 9. Guardar el libro de trabajo en un flujo de memoria
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ------------ Aplicación Flask ------------
app = Flask(__name__)

# ** NUEVA RUTA PARA MOSTRAR EL FORMULARIO **
@app.get('/')
def index():
    return render_template('index.html')

@app.get('/excel')
def descargar_excel():
    ini = request.args.get('ini')
    fin = request.args.get('fin')
    if not ini or not fin:
        return jsonify({'error': 'Parámetros ini y fin requeridos (YYYY-MM-DD)'}), 400
    try:
        d_ini = datetime.strptime(ini, '%Y-%m-%d')
        d_fin = datetime.strptime(fin, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    if d_fin < d_ini:
        return jsonify({'error': 'fin debe ser posterior a ini'}), 400

    stream = generar_excel(d_ini, d_fin)
    return send_file(
        stream,
        as_attachment=True,
        download_name='programacion.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)


